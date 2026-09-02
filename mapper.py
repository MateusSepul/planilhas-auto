"""
mapper.py — Lógica de leitura, mapeamento e escrita de planilhas Excel.
"""

import difflib
import re
import unicodedata
from typing import Callable, Dict, List, Optional, Set, Tuple
import openpyxl


SYNONYM_GROUPS = [
    {"nome", "cliente", "razao", "razaosocial", "titular", "nomecliente", "nomecompleto", "usuario", "comprador", "socio", "fornecedor"},
    {"email", "correioeletronico", "contato", "mail", "emailcontato"},
    {"telefone", "tel", "celular", "cel", "fone", "whatsapp", "contatotel", "telefonecontato"},
    {"valor", "total", "preco", "valortotal", "quantia", "montante", "saldo", "vlr", "vlrtotal", "precototal", "custo", "subtotal"},
    {"data", "dt", "dataemissao", "datavencimento", "ref", "referencia", "periodo", "datacadastro", "dia"},
    {"cpf", "cnpj", "documento", "doc", "cpfcnpj", "nrodocumento", "numdocumento"},
    {"endereco", "logradouro", "rua", "localizacao", "enderecocompleto", "bairro"},
    {"cidade", "municipio"},
    {"estado", "uf"},
    {"cep", "codigopostal", "zip", "zipcode"},
    {"codigo", "cod", "id", "identificador", "sku", "chave", "numero", "num", "nro"},
    {"descricao", "desc", "historico", "produto", "item", "detalhe", "detalhes", "observacao", "obs", "especificacao"},
    {"quantidade", "qtd", "quant", "volume", "unidades", "itens", "estoque", "quantitativo"},
    {"status", "situacao", "estadoatual", "fase"},
]


def _normalize_text(text: str) -> str:
    """Remove acentos, converte para minúsculas e remove espaços nas pontas."""
    if not text:
        return ""
    text = str(text).strip().lower()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("utf-8")
    return text


def _tokenize(text: str) -> Set[str]:
    """Divide a string em tokens alfanuméricos."""
    norm = _normalize_text(text)
    return set(re.findall(r"[a-z0-9]+", norm))


def _clean_str(text: str) -> str:
    """Remove todos os caracteres não alfanuméricos."""
    norm = _normalize_text(text)
    return re.sub(r"[^a-z0-9]", "", norm)


def _get_synonym_group(token: str) -> Optional[int]:
    """Retorna o índice do grupo de sinônimos se o token pertencer a algum."""
    for idx, grp in enumerate(SYNONYM_GROUPS):
        if token in grp:
            return idx
    return None


def _score_match(src: str, dst: str) -> float:
    """Calcula a pontuação de similaridade/compatibilidade entre duas colunas (0.0 a 1.0)."""
    s_clean = _clean_str(src)
    d_clean = _clean_str(dst)
    if not s_clean or not d_clean:
        return 0.0

    # Correspondência exata normalizada
    if s_clean == d_clean:
        return 1.0

    s_tokens = _tokenize(src)
    d_tokens = _tokenize(dst)

    # Interseção de tokens exatos
    common_tokens = s_tokens.intersection(d_tokens)
    if common_tokens:
        overlap = len(common_tokens) / max(len(s_tokens), len(d_tokens))
        return 0.85 + (overlap * 0.1)

    # Grupo de sinônimos no texto limpo completo
    s_syn_group = _get_synonym_group(s_clean)
    d_syn_group = _get_synonym_group(d_clean)
    if s_syn_group is not None and s_syn_group == d_syn_group:
        return 0.90

    # Grupo de sinônimos por tokens
    for st in s_tokens:
        st_grp = _get_synonym_group(st)
        if st_grp is None:
            continue
        for dt in d_tokens:
            dt_grp = _get_synonym_group(dt)
            if dt_grp is not None and st_grp == dt_grp:
                return 0.82

    # Inclusão de substring para termos significativos (>= 3 caracteres)
    if len(s_clean) >= 3 and len(d_clean) >= 3:
        if d_clean in s_clean or s_clean in d_clean:
            ratio = len(min(d_clean, s_clean, key=len)) / len(max(d_clean, s_clean, key=len))
            return 0.75 + (ratio * 0.15)

    # Similaridade difflib
    return difflib.SequenceMatcher(None, s_clean, d_clean).ratio()


def auto_match_columns(
    source_cols: List[str],
    dest_cols: List[str],
    similarity_threshold: float = 0.65
) -> Dict[str, str]:
    """
    Relaciona automaticamente colunas de origem e destino calculando
    compatibilidade por igualdade, sinônimos, palavras-chave e similaridade.

    Args:
        source_cols: Lista de colunas da planilha de origem
        dest_cols: Lista de colunas da planilha de destino
        similarity_threshold: Limiar mínimo para aceitar um relacionamento

    Returns:
        Dict mapeando {coluna_destino: coluna_origem}
    """
    candidates = []
    for d in dest_cols:
        if not d or not str(d).strip():
            continue
        for s in source_cols:
            if not s or not str(s).strip():
                continue
            sc = _score_match(s, d)
            if sc >= similarity_threshold:
                candidates.append((sc, d, s))

    # Ordena por pontuação decrescente para priorizar as melhores correspondências
    candidates.sort(key=lambda x: x[0], reverse=True)

    matches: Dict[str, str] = {}
    matched_dest: Set[str] = set()
    used_sources: Set[str] = set()

    for sc, d, s in candidates:
        if d in matched_dest or s in used_sources:
            continue
        matches[d] = s
        matched_dest.add(d)
        used_sources.add(s)

    return matches



def get_columns(filepath: str, sheet_index: int = 0) -> Tuple[List[str], str]:
    """
    Lê os cabeçalhos (primeira linha) de uma planilha Excel.

    Args:
        filepath: Caminho para o arquivo .xlsx
        sheet_index: Índice da aba (0 = primeira aba)

    Returns:
        Tupla (lista_de_colunas, nome_da_aba)
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    ws = wb[sheet_names[sheet_index]]
    headers = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for cell in row:
            headers.append(str(cell) if cell is not None else "")
    wb.close()
    return headers, sheet_names[sheet_index]


def get_sheet_names(filepath: str) -> List[str]:
    """Retorna os nomes de todas as abas de uma planilha."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def execute_mapping(
    source_path: str,
    dest_path: str,
    output_path: str,
    mapping: List[Dict],
    source_sheet: int = 0,
    dest_sheet: int = 0,
    import_formulas: bool = False,
    progress_callback: Optional[Callable[[int, int], None]] = None,
) -> int:
    """
    Aplica o mapeamento de colunas: copia dados da planilha de origem
    para a planilha de destino e salva em output_path.

    Args:
        source_path: Caminho da planilha de origem (fonte dos dados)
        dest_path: Caminho da planilha de destino (template)
        output_path: Onde salvar o arquivo preenchido
        mapping: Lista de {'source': 'NomeColOrigem', 'destination': 'NomeColDestino'}
        source_sheet: Índice da aba na planilha de origem
        dest_sheet: Índice da aba na planilha de destino
        import_formulas: Se True, preserva e importa as fórmulas/funções do Excel em vez de apenas os valores calculados
        progress_callback: Função chamada com (linha_atual, total_linhas)

    Returns:
        Número de linhas copiadas.
    """
    # Lê origem (se import_formulas for True, data_only=False preserva as funções/fórmulas)
    wb_src = openpyxl.load_workbook(source_path, data_only=not import_formulas)
    ws_src = wb_src[wb_src.sheetnames[source_sheet]]

    # Lê destino (template)
    wb_dst = openpyxl.load_workbook(dest_path)
    ws_dst = wb_dst[wb_dst.sheetnames[dest_sheet]]

    # Mapeia nomes de colunas para índices na origem
    src_headers = [str(c.value) if c.value is not None else "" for c in ws_src[1]]
    dst_headers = [str(c.value) if c.value is not None else "" for c in ws_dst[1]]

    # Cria dicionário: coluna_destino -> coluna_origem (índice 1-based)
    col_map: Dict[int, int] = {}  # dst_col_idx -> src_col_idx
    for rule in mapping:
        src_name = rule.get("source", "")
        dst_name = rule.get("destination", "")
        if src_name in src_headers and dst_name in dst_headers:
            src_idx = src_headers.index(src_name) + 1
            dst_idx = dst_headers.index(dst_name) + 1
            col_map[dst_idx] = src_idx

    # Conta linhas de dados na origem (sem o cabeçalho)
    src_rows = list(ws_src.iter_rows(min_row=2, values_only=True))
    total = len(src_rows)

    # Determina a primeira linha de dados no destino
    # Preserva as linhas de cabeçalho do template (linha 1)
    dst_data_start = 2

    # Limpa dados existentes no destino a partir da linha de dados
    # (mantém o cabeçalho intacto)
    if ws_dst.max_row >= dst_data_start:
        for row in ws_dst.iter_rows(min_row=dst_data_start):
            for cell in row:
                cell.value = None

    # Copia dados linha a linha
    for i, src_row_values in enumerate(src_rows):
        dst_row_num = dst_data_start + i

        for dst_col_idx, src_col_idx in col_map.items():
            # src_row_values é 0-indexed
            value = src_row_values[src_col_idx - 1]
            ws_dst.cell(row=dst_row_num, column=dst_col_idx, value=value)

        if progress_callback:
            progress_callback(i + 1, total)

    wb_dst.save(output_path)
    wb_src.close()

    return total
